"""
BOMMEL - BOM Merger for Excel Lists
Python port of bommel.bas

Author: Daniel Leidner
License: MIT

Usage:
    python bommel.py [old_file] [new_file] [-o output_file]

Install dependency first:
    pip install openpyxl
"""

import sys
import time
import argparse
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.formatting.rule import FormulaRule
    from openpyxl.formula.translate import Translator
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

# =============================================================================
# CONFIGURATION
# =============================================================================

OLD_FILE = "POD.02_20260416.xlsx"
NEW_FILE  = "POD.02_20260414.xlsx"

TRANSFER_COLS = {
    "Status", "Angebotsnummer", "Bemerkung", "Lieferzeit",
    "Liefertermin", "Link", "Preis Gesamt", "Lieferant",
}

FILE_COL              = "Datei"
ASSEMBLY_COL          = "Baugruppe 2"
CATEGORY_COL          = "Kategorie"
MATCHING_ASSEMBLY_COLS = ["Baugruppe 1", "Baugruppe 2"]

CRITICAL_STATUSES = {"ordered", "paid", "delivered", "completed"}

QUANTITY_CHANGED_COLOR = "FFB3B3"

# =============================================================================
# HELPERS
# =============================================================================

def solid_fill(hex_color: str) -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=hex_color)


def read_sheet(path: Path):
    """Return (headers: list[str], rows: list[list], sheet_name: str)."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.worksheets[0]
    data = [[cell.value for cell in row] for row in ws.iter_rows()]
    if not data:
        return [], [], ws.title
    headers = [str(h) if h is not None else "" for h in data[0]]
    return headers, data[1:], ws.title


def merge_headers(new_headers: list[str], old_headers: list[str]) -> list[str]:
    """
    Start with new_headers order. For each column in old_headers missing from
    result, insert it after its nearest left neighbour that exists in result,
    or at position 0 if no neighbour found.
    """
    result = list(new_headers)
    result_set = set(result)

    for i, col in enumerate(old_headers):
        if not col or col in result_set:
            continue
        left_neighbor = next(
            (old_headers[j] for j in range(i - 1, -1, -1) if old_headers[j] in result_set),
            None,
        )
        if left_neighbor is not None:
            insert_pos = result.index(left_neighbor) + 1
            print(f"   + '{col}' inserted after '{left_neighbor}'")
        else:
            insert_pos = 0
            print(f"   + '{col}' inserted at beginning")
        result.insert(insert_pos, col)
        result_set.add(col)

    return result


def make_key(row: list, header_idx: dict, assembly_cols: list[str]) -> str:
    datei = str(row[header_idx[FILE_COL]] or "").strip() if FILE_COL in header_idx else ""
    parts = [datei or "NO_FILE"]
    for col in assembly_cols:
        val = str(row[header_idx[col]] or "").strip() if col in header_idx else ""
        parts.append(val)
    return "|".join(parts)


def read_formulas(path: Path) -> list[list]:
    """Load with data_only=False to get formula strings instead of cached values."""
    wb = openpyxl.load_workbook(path, data_only=False)
    ws = wb.worksheets[0]
    return [[cell.value for cell in row] for row in ws.iter_rows()]


def col_widths(ws):
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        width = max((len(str(c.value or "")) for c in col_cells), default=0)
        ws.column_dimensions[col_letter].width = min(max(width + 2, 8), 60)


# =============================================================================
# MAIN
# =============================================================================

def merge(old_file: str, new_file: str, output_file: str | None = None) -> str:
    start = time.time()
    old_path, new_path = Path(old_file), Path(new_file)

    for p in (old_path, new_path):
        if not p.exists():
            print(f"ERROR: '{p}' not found.")
            sys.exit(1)

    print("=" * 80)
    print("BOMMEL - PROCUREMENT LIST MERGER")
    print("=" * 80)

    # 1. Read ----------------------------------------------------------------
    print("\n1. Reading data...")
    old_headers, old_rows, old_sheet = read_sheet(old_path)
    new_headers, new_rows, new_sheet = read_sheet(new_path)
    new_formula_rows = read_formulas(new_path)[1:]  # skip header row
    print(f"   Old: {old_path.name} / {old_sheet}  →  {len(old_rows)} rows")
    print(f"   New: {new_path.name} / {new_sheet}  →  {len(new_rows)} rows")

    # 2. Column mapping ------------------------------------------------------
    print("\n2. Creating column mapping...")
    all_headers = merge_headers(new_headers, old_headers)
    all_headers += ["_status_flag", "_quantity_changed"]

    old_hi = {h: i for i, h in enumerate(old_headers)}
    new_hi = {h: i for i, h in enumerate(new_headers)}
    out_hi = {h: i for i, h in enumerate(all_headers)}

    # 3. Matching keys -------------------------------------------------------
    old_keys = [make_key(r, old_hi, MATCHING_ASSEMBLY_COLS) for r in old_rows]
    new_keys = [make_key(r, new_hi, MATCHING_ASSEMBLY_COLS) for r in new_rows]

    old_key_map: dict[str, int] = {}
    for i, k in enumerate(old_keys):
        if k not in old_key_map:
            old_key_map[k] = i

    # 4. Build output rows ---------------------------------------------------
    print("\n3. Matching and transferring data...")

    status_out  = out_hi.get("Status", -1)
    flag_out    = out_hi["_status_flag"]
    qty_out     = out_hi["_quantity_changed"]
    bg2_out     = out_hi.get(ASSEMBLY_COL, -1)
    kat_out     = out_hi.get(CATEGORY_COL, -1)
    anz_new     = new_hi.get("Anzahl", -1)
    anz_old     = old_hi.get("Anzahl", -1)
    anzg_new    = new_hi.get("Anzahl gesamt", -1)
    anzg_old    = old_hi.get("Anzahl gesamt", -1)
    status_old  = old_hi.get("Status", -1)

    data_cols = all_headers[:-2]  # exclude _status_flag, _quantity_changed

    def try_float(v):
        try:
            return float(v) if v is not None else None
        except (ValueError, TypeError):
            return None

    matched = updated = new_cnt = qty_warnings = 0
    output_rows: list[list] = []

    for i, new_row in enumerate(new_rows):
        out = [None] * len(all_headers)
        status_flag = ""
        qty_changed = False
        has_changes = False

        old_i = old_key_map.get(new_keys[i], -1)

        if old_i >= 0:
            matched += 1
            old_row = old_rows[old_i]

            # Quantity change detection
            for ni, oi in ((anz_new, anz_old), (anzg_new, anzg_old)):
                if not qty_changed and ni >= 0 and oi >= 0:
                    ov, nv = try_float(old_row[oi]), try_float(new_row[ni])
                    if ov is not None and nv is not None and ov != nv:
                        qty_changed = True

            # Critical status check
            if qty_changed and status_old >= 0:
                if str(old_row[status_old] or "").strip().lower() in CRITICAL_STATUSES:
                    qty_warnings += 1

            # Fill output columns
            for j, col in enumerate(data_cols):
                oi = old_hi.get(col, -1)
                ni = new_hi.get(col, -1)
                if col in TRANSFER_COLS and oi >= 0:
                    old_val = old_row[oi]
                    if old_val is not None and str(old_val).strip():
                        out[j] = old_val
                        has_changes = True
                    elif ni >= 0:
                        out[j] = new_row[ni]
                elif ni >= 0:
                    out[j] = new_row[ni]

            if has_changes:
                status_flag = "UPDATED"
                updated += 1
        else:
            new_cnt += 1
            status_flag = "NEW"
            for j, col in enumerate(data_cols):
                ni = new_hi.get(col, -1)
                if ni >= 0:
                    out[j] = new_row[ni]
                    if col == "Status" and not str(out[j] or "").strip():
                        out[j] = "new"

        # Empty Status → "-"
        if status_out >= 0 and not str(out[status_out] or "").strip():
            out[status_out] = "-"

        out[flag_out] = status_flag
        out[qty_out]  = qty_changed
        output_rows.append(out)

    # 5. Deleted rows --------------------------------------------------------
    print("\n4. Finding deleted entries...")
    new_key_set = set(new_keys)
    deleted = [old_rows[i] for i, k in enumerate(old_keys) if k not in new_key_set]
    print(f"   Found: {len(deleted)} deleted entries")

    # 6. Write workbook ------------------------------------------------------
    print("\n5. Writing output workbook...")
    wb = openpyxl.Workbook()

    # --- Main_List ----------------------------------------------------------
    ws = wb.active
    ws.title = "Main_List"

    ws.append(all_headers)
    for col in range(1, len(all_headers) + 1):
        c = ws.cell(1, col)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = solid_fill("404040")

    for row_idx, row in enumerate(output_rows, start=2):
        ws.append(row)

        bg2_val = str(row[bg2_out] or "").strip() if bg2_out >= 0 else "x"
        kat_val = str(row[kat_out] or "").strip().lower() if kat_out >= 0 else ""

        # Baugruppe header rows: direct formatting applied last so it is never
        # overridden by a matching CF rule (their Status is "-" / qty_changed is
        # False, so no CF rule fires on them anyway).
        if bg2_val == "" and kat_val == "baugruppe":
            for col in range(1, len(all_headers) + 1):
                c = ws.cell(row_idx, col)
                c.fill = solid_fill("A5A5A5")
                c.font = Font(bold=True, color="FFFFFF")

    # Second pass: restore formulas from new workbook (overwrites cached values).
    # Uses Translator to adjust column references when old-only columns were
    # inserted before a formula's referenced columns.
    formula_count = 0
    for new_i, form_row in enumerate(new_formula_rows):
        out_row = new_i + 2          # +1 header, +1 because enumerate is 0-based
        src_wb_row = new_i + 2       # same row number in new workbook
        for src_col_idx, col_name in enumerate(new_headers):
            if col_name in TRANSFER_COLS or col_name not in out_hi:
                continue
            if src_col_idx >= len(form_row):
                continue
            formula_val = form_row[src_col_idx]
            if not isinstance(formula_val, str) or not formula_val.startswith("="):
                continue
            out_col_idx = out_hi[col_name]
            src_ref = f"{get_column_letter(src_col_idx + 1)}{src_wb_row}"
            dst_ref = f"{get_column_letter(out_col_idx + 1)}{out_row}"
            try:
                translated = Translator(formula_val, origin=src_ref).translate_formula(dst_ref)
                ws.cell(out_row, out_col_idx + 1).value = translated
                formula_count += 1
            except Exception:
                pass  # keep computed value on translation failure
    if formula_count:
        print(f"   ✓ {formula_count} formulas restored from new workbook")

    # Conditional formatting — rules fire in priority order (first = highest).
    # Colors update automatically whenever the user changes the Status dropdown.
    last_row  = len(output_rows) + 1
    cf_range  = f"A2:{get_column_letter(len(all_headers))}{last_row}"
    sc = f"${get_column_letter(status_out + 1)}2" if status_out >= 0 else None
    qc = f"${get_column_letter(qty_out + 1)}2"

    cf_rules = [
        FormulaRule(formula=[f"={qc}=TRUE"],
                    fill=solid_fill(QUANTITY_CHANGED_COLOR), font=Font(bold=True), stopIfTrue=True),
    ]
    if sc:
        cf_rules += [
            FormulaRule(formula=[f'={sc}="new"'],
                        fill=solid_fill("FFCCB3"), stopIfTrue=True),
            FormulaRule(formula=[f'={sc}="requested"'],
                        fill=solid_fill("FFFFCC"), stopIfTrue=True),
            FormulaRule(formula=[f'=OR({sc}="offered",{sc}="angeboten")'],
                        fill=solid_fill("DDCCFF"), stopIfTrue=True),
            FormulaRule(formula=[f'={sc}="ordered"'],
                        fill=solid_fill("ADD8E6"), stopIfTrue=True),
            FormulaRule(formula=[f'={sc}="paid"'],
                        fill=solid_fill("C6EFCE"), stopIfTrue=True),
            FormulaRule(formula=[f'={sc}="delivered"'],
                        fill=solid_fill("93C47D"), stopIfTrue=True),
            FormulaRule(formula=[f'={sc}="completed"'],
                        fill=solid_fill("6AA84F"), font=Font(bold=True), stopIfTrue=True),
            FormulaRule(formula=[f'={sc}="postponed"'],
                        fill=solid_fill("D9D9D9"), font=Font(italic=True), stopIfTrue=True),
        ]

    for rule in cf_rules:
        ws.conditional_formatting.add(cf_range, rule)

    # AutoFilter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(all_headers))}1"

    # Status dropdown
    if status_out >= 0:
        sl = get_column_letter(status_out + 1)
        dv = DataValidation(
            type="list",
            formula1='"-,new,requested,offered,ordered,paid,delivered,completed,postponed"',
            allow_blank=True,
            showDropDown=False,
        )
        dv.sqref = f"{sl}2:{sl}{len(output_rows) + 1}"
        ws.add_data_validation(dv)

    col_widths(ws)

    # --- Deleted_Entries ----------------------------------------------------
    if deleted:
        dws = wb.create_sheet("Deleted_Entries")
        dws.append(old_headers)
        for col in range(1, len(old_headers) + 1):
            c = dws.cell(1, col)
            c.font = Font(bold=True)
            c.fill = solid_fill("D3D3D3")
        for row in deleted:
            dws.append(row)
            ri = dws.max_row
            for col in range(1, len(old_headers) + 1):
                dws.cell(ri, col).fill = solid_fill("FFB6C1")
        col_widths(dws)

    # --- Log ----------------------------------------------------------------
    lws = wb.create_sheet("Log")
    elapsed = round(time.time() - start, 2)

    lws["A1"] = "PROCUREMENT LIST MERGE LOG"
    lws["A1"].font = Font(bold=True, size=14)
    lws["A3"], lws["B3"] = "Date:",     datetime.now().strftime("%d.%m.%Y %H:%M:%S")
    lws["A4"], lws["B4"] = "Old List:", f"{old_path.name} / {old_sheet}"
    lws["A5"], lws["B5"] = "New List:", f"{new_path.name} / {new_sheet}"
    lws["A7"] = "STATISTICS"
    lws["A7"].font = Font(bold=True)
    lws["A8"],  lws["B8"]  = "Matched Rows:",      matched
    lws["A9"],  lws["B9"]  = "Of which updated:",  updated
    lws["A10"], lws["B10"] = "New Entries:",        new_cnt
    lws["A11"], lws["B11"] = "Deleted Entries:",    len(deleted)
    if qty_warnings:
        lws["A12"] = "⚠️ Quantity Changes (ordered):"
        lws["B12"] = qty_warnings
        lws["A12"].font = lws["B12"].font = Font(bold=True)
    lws["A13"], lws["B13"] = "Processing Time:", f"{elapsed}s"

    lws["A15"] = "COLOR LEGEND (by Status)"
    lws["A15"].font = Font(bold=True)
    legend = [
        ("⚪ White",        "not yet touched / -", None,               False, False),
        ("🔴 Light Red",    "_quantity_changed",   QUANTITY_CHANGED_COLOR, True, False),
        ("🟠 Coral",        "new",                 "FFCCB3",           False, False),
        ("🟡 Yellow",       "requested",           "FFFFCC",           False, False),
        ("🟣 Light Purple", "offered",             "DDCCFF",           False, False),
        ("🔵 Light Blue",   "ordered",             "ADD8E6",           False, False),
        ("🟢 Light Green",  "paid",                "C6EFCE",           False, False),
        ("🟢 Medium Green", "delivered",           "93C47D",           False, False),
        ("🟢 Dark Green",   "completed",           "6AA84F",           True,  False),
        ("⚫ Gray",         "postponed",           "D9D9D9",           False, True),
    ]
    for r, (label, desc, color, bold, italic) in enumerate(legend, start=16):
        ca = lws.cell(r, 1, label)
        lws.cell(r, 2, desc)
        if color:
            ca.fill = solid_fill(color)
        if bold or italic:
            ca.font = Font(bold=bold, italic=italic)

    lws.column_dimensions["A"].width = 22
    lws.column_dimensions["B"].width = 32

    # Save -------------------------------------------------------------------
    if output_file is None:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = str(old_path.parent / f"BOMMEL_merged_{ts}.xlsx")

    wb.save(output_file)

    print("\n" + "=" * 80)
    print("DONE!")
    print("=" * 80)
    print(f"Matched Rows:     {matched}")
    print(f"Of which updated: {updated}")
    print(f"New Entries:      {new_cnt}")
    print(f"Deleted Entries:  {len(deleted)}")
    if qty_warnings:
        print(f"\n⚠️  WARNING: {qty_warnings} quantity changes in ordered parts!")
        print("   → These rows are marked LIGHT RED — manual review needed")
    print(f"\nProcessing Time:  {elapsed}s")
    print(f"✅ Output saved:  {output_file}")
    return output_file


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="BOMMEL - BOM Merger for Excel Lists")
    parser.add_argument("old_file", nargs="?", default=OLD_FILE,
                        help=f"Old workbook (default: {OLD_FILE})")
    parser.add_argument("new_file", nargs="?", default=NEW_FILE,
                        help=f"New workbook (default: {NEW_FILE})")
    parser.add_argument("-o", "--output", default=None,
                        help="Output filename (default: BOMMEL_merged_<timestamp>.xlsx)")
    args = parser.parse_args()
    merge(args.old_file, args.new_file, args.output)
